"""
Nucleo compartido: logica de combinar y extraer PDFs.

Tanto los scripts de linea de comandos (combinar_pdfs.py / extraer_pdfs.py)
como la interfaz web (app.py) usan estas funciones. Asi la logica vive en un
solo lugar y no se duplica.

Cada funcion NO imprime nada: devuelve un diccionario con el resultado, y quien
la llama decide como mostrarlo (consola o navegador).

Ver CLAUDE.md para el contexto completo y las decisiones de diseno.
"""

import zipfile
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfWriter


# --------------------------------------------------------------------------- #
# COMBINAR
# --------------------------------------------------------------------------- #

# Nombres de encabezado aceptados para la columna de control (comparacion en
# minusculas). Si el Excel no tiene ninguna, se crea una llamada "Estado".
NOMBRES_ESTADO = {"estado", "combinado", "ya combinado", "status", "creado"}


def _claves_archivo(stem):
    """Claves candidatas de folio para un nombre de archivo (en MAYUSCULAS).

    Cubre los formatos reales de la carpeta de facturas:
      - "10029998.pdf" / "ACRED3340.pdf" / "GDL1A-21858.pdf" -> nombre completo
      - "GPA8402219Y1_FA_10322157_FOCA000109KD3.pdf"         -> cada segmento "_"
        (asi el folio 10322157 casa aunque venga rodeado de RFC/tipo)

    No se parte por guion "-": es parte de folios como GDL1A-21858.
    Mayusculas para casar sin importar mayus/minus.
    """
    s = stem.upper()
    claves = set(s.split("_"))
    claves.add(s)
    return {k.strip() for k in claves if k.strip()}


def construir_indice(carpeta_pdfs):
    """Recorre la carpeta UNA vez. Devuelve (indice {clave: ruta}, multiples).

    'multiples' = claves presentes en mas de un archivo (posibles duplicados).
    Solo se avisan las que de verdad se usan al combinar (ver combinar()).
    """
    indice = {}
    multiples = set()
    for p in Path(carpeta_pdfs).glob("*.pdf"):
        for clave in _claves_archivo(p.stem):
            if clave in indice and indice[clave] != p:
                multiples.add(clave)
            indice[clave] = p
    return indice, multiples


def _celda_str(v):
    """Normaliza el valor de una celda a texto limpio (sin '.0' en enteros)."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    return str(v).strip()


def combinar(excel_path, carpeta_pdfs, carpeta_salida, tiene_encabezado=True):
    """Combina PDFs por fila del Excel, con CONTROL de duplicados.

    Reglas del Excel:
      - Columna de CONTROL: la que se llame Estado/Combinado/etc. (si no existe
        y hay encabezado, se crea una llamada 'Estado').
      - De las columnas restantes: la ULTIMA es el nombre del archivo de salida;
        las demas son los folios a combinar (cantidad variable).

    Control de duplicados:
      - Si la celda de Estado de una fila YA tiene algo, esa fila se OMITE (no
        se vuelve a generar). Asi el Excel sirve de bitacora.
      - Al generar un PDF, se escribe 'Creado AAAA-MM-DD HH:MM' en su Estado y se
        guarda el Excel. (Requiere que el Excel NO este abierto en otro programa.)
      - Las filas con folios faltantes NO se marcan (se reintentan a la proxima).

    Devuelve dict:
        {
          "ok": int,                 # PDFs generados en esta corrida
          "generados": [str],        # nombres creados en esta corrida
          "omitidos": int,           # filas saltadas por ya estar creadas
          "errores": [str],          # filas con problemas (no se generaron)
          "duplicados": [str],       # folios repetidos en la carpeta origen
          "control_activo": bool,    # True si se uso la columna Estado
        }
    """
    out = Path(carpeta_salida)
    out.mkdir(parents=True, exist_ok=True)

    indice, multiples = construir_indice(carpeta_pdfs)
    duplicados_usados = set()

    wb = load_workbook(excel_path)
    ws = wb.active

    # --- Encabezados y fila donde empiezan los datos ---
    encabezados = [c.value for c in ws[1]] if ws.max_row >= 1 else []
    # quitar columnas vacias al final
    while encabezados and encabezados[-1] in (None, ""):
        encabezados.pop()
    n_cols = len(encabezados)
    primera_fila_datos = 2 if tiene_encabezado else 1

    # --- Localizar (o crear) la columna de Estado ---
    col_estado = None  # 1-based
    control_activo = False
    if tiene_encabezado and n_cols:
        for i, h in enumerate(encabezados):
            if str(h).strip().lower() in NOMBRES_ESTADO:
                col_estado = i + 1
                break
        if col_estado is None:
            # crearla al final
            col_estado = n_cols + 1
            ws.cell(row=1, column=col_estado, value="Estado")
            n_cols = col_estado
        control_activo = True

    # columnas de datos = todas menos la de estado; ultima = nombre, resto folios
    cols_datos = [c for c in range(1, n_cols + 1) if c != col_estado]
    if not cols_datos:
        wb.close()
        return {"ok": 0, "generados": [], "omitidos": 0,
                "errores": ["El Excel no tiene columnas de datos."],
                "duplicados": [], "control_activo": control_activo}
    col_nombre = cols_datos[-1]
    cols_folios = cols_datos[:-1]

    marca = datetime.now().strftime("Creado %Y-%m-%d %H:%M")
    errores = []
    generados = []
    omitidos = 0
    hubo_cambios = False

    for r in range(primera_fila_datos, ws.max_row + 1):
        # fila totalmente vacia -> ignorar
        if all(ws.cell(row=r, column=c).value in (None, "") for c in range(1, n_cols + 1)):
            continue

        # --- Control: si ya fue creada, omitir ---
        if col_estado is not None:
            estado_val = _celda_str(ws.cell(row=r, column=col_estado).value)
            if estado_val != "":
                omitidos += 1
                continue

        nombre_final = _celda_str(ws.cell(row=r, column=col_nombre).value)
        folios = [_celda_str(ws.cell(row=r, column=c).value) for c in cols_folios]
        folios = [f for f in folios if f != ""]

        if not folios:
            errores.append(f"Fila {r}: sin folios")
            continue
        if nombre_final == "":
            errores.append(f"Fila {r}: sin nombre final")
            continue

        w = PdfWriter()
        faltantes = []
        for folio in folios:
            clave = folio.upper()
            ruta = indice.get(clave)
            if ruta:
                w.append(str(ruta))
                if clave in multiples:
                    duplicados_usados.add(folio)
            else:
                faltantes.append(folio)

        # NUNCA generar un PDF combinado incompleto (regla de control).
        if faltantes:
            errores.append(
                f"Fila {r}: no se encontro folio(s) {', '.join(faltantes)}")
            w.close()
            continue

        nombre_archivo = f"{nombre_final}.pdf"
        with open(out / nombre_archivo, "wb") as f:
            w.write(f)
        w.close()
        generados.append(nombre_archivo)

        # marcar como creada
        if col_estado is not None:
            ws.cell(row=r, column=col_estado, value=marca)
            hubo_cambios = True

    # Guardar el Excel solo si marcamos algo nuevo.
    if hubo_cambios:
        try:
            wb.save(excel_path)
        except PermissionError:
            wb.close()
            raise PermissionError(
                "No se pudo guardar el Excel (¿lo tienes abierto?). "
                "Cierralo y vuelve a intentar.")
    wb.close()

    return {
        "ok": len(generados),
        "generados": generados,
        "omitidos": omitidos,
        "errores": errores,
        "duplicados": sorted(duplicados_usados),
        "control_activo": control_activo,
    }


# --------------------------------------------------------------------------- #
# EXTRAER
# --------------------------------------------------------------------------- #

def _nombre_unico(destino):
    """Devuelve una ruta que no existe (renombra duplicados: factura_1.pdf...)."""
    if not destino.exists():
        return destino
    base, ext, n = destino.stem, destino.suffix, 1
    while True:
        nuevo = destino.with_name(f"{base}_{n}{ext}")
        if not nuevo.exists():
            return nuevo
        n += 1


def extraer_zips(carpeta_zips, carpeta_salida, aplanar=True):
    """Extrae todos los PDFs contenidos en los .zip de la carpeta.

    Devuelve dict:
        {
          "total": int,        # PDFs extraidos
          "num_zips": int,     # ZIPs procesados
          "errores": [str],    # ZIPs con problema (no detiene el proceso)
        }
    """
    src = Path(carpeta_zips)
    out = Path(carpeta_salida)
    out.mkdir(parents=True, exist_ok=True)

    total_pdfs = 0
    errores = []
    zips = list(src.glob("*.zip"))

    for z in zips:
        try:
            with zipfile.ZipFile(z) as archivo:
                pdfs = [f for f in archivo.namelist()
                        if f.lower().endswith(".pdf") and not f.endswith("/")]
                for f in pdfs:
                    solo_nombre = Path(f).name
                    if aplanar:
                        destino = _nombre_unico(out / solo_nombre)
                    else:
                        sub = out / z.stem
                        sub.mkdir(exist_ok=True)
                        destino = _nombre_unico(sub / solo_nombre)
                    with archivo.open(f) as origen, open(destino, "wb") as salida:
                        salida.write(origen.read())
                    total_pdfs += 1
        except zipfile.BadZipFile:
            errores.append(f"{z.name}: archivo corrupto o no es ZIP valido")
        except Exception as e:
            errores.append(f"{z.name}: {e}")

    return {
        "total": total_pdfs,
        "num_zips": len(zips),
        "errores": errores,
    }
